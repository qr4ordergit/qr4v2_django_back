from django.shortcuts import render
from rest_framework.views import APIView
from django.utils import translation
from django.utils.translation import gettext as _
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
import pandas as pd
import polib
import os



class LanguageDetails(APIView):

    def checkLang(self,lang):
        if lang in[code for code, _ in settings.LANGUAGES]:
            return True
        return False

    def get(self,request,lang):

        if not self.checkLang(lang):
            return JsonResponse({'message':"Language not present in the system"})
        
        translation.activate(lang)
        translations = {}

        po = polib.pofile(f"locale/{lang}/LC_MESSAGES/django.po")

        for entry in po:
            translations[str(entry.msgid)] = _(str(entry.msgid))
        return JsonResponse(translations)

        
    def export(self,request,lang):
        fill = PatternFill(start_color="5983B0", end_color="5983B0", fill_type="solid")
        header_font = Font(bold=True)

        msgid_list = []
        msgstr_list = []

        po = polib.pofile(f"locale/{lang}/LC_MESSAGES/django.po")

        for entry in po:
            msgid_list.append(entry.msgid)
            msgstr_list.append(entry.msgstr)

        data = {'Code': msgid_list, 'Phrase': msgstr_list}
        df = pd.DataFrame(data)
        wb = Workbook()
        ws = wb.active

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill

        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = cell.alignment.copy(wrap_text=True)

        max_length = 0
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
        ws.column_dimensions['B'].width = max_length + 2

        wb.save('your_existing_workbook.xlsx')

        # excel_file_path = 'path/to/output/file.xlsx'

        print(df)
    
    def post(self,request,lang):
        if not self.checkLang(lang):
            return JsonResponse({'message':"Language not present in the system"})
        
        res = self.export(request,lang)

        # file_path = os.path.join(settings.MEDIA_ROOT, lang)
        # print(file_path)
        # if os.path.exists(file_path):
        #     with open(file_path, 'w') as file:
        #         response = HttpResponse(file.write(), content_type='application/vnd.ms-excel')
        #         response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        #         return response
        # else:
        return JsonResponse({"message": "Report not found."})            


